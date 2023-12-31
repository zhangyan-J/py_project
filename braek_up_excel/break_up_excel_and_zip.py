# !/usr/bin/env python
# -*- coding:utf-8 -*-

# 先把一个excel里的每个sheet文件按照sheetname分成单个excel文件，按文件后缀名为xlsx放入一个文件夹并且打包成zip
import datetime

from openpyxl import load_workbook
from openpyxl import Workbook
import os
import shutil
import zipfile

# 修改工作目录至对应的文件夹，拆分后的文件会保存到此目录
# 临时修改工作路径，作为拆分后的文件存储的地方
os.chdir(input("请输入存放的路径: "))
# os.chdir(r"D:\Practice\分割文件")
# 导入需要拆分的excel，赋值给wb
a = input("请输入要拆分的excel: ")
# print(a)
wb = load_workbook(a)
# wb = load_workbook('D:\project\py_project\\braek_up_excel\case.xlsx')
# wb = load_workbook(r'D:\Practice\分割文件\5.0适配func.xlsx')
sheetnames = wb.sheetnames
# 得到excel中的每个sheet名
for name in sheetnames:
    ws = wb.get_sheet_by_name(name)
    # print(ws)

    # 创建新的excel
    wb2 = Workbook()
    # 获取当前sheet，调用正在运行的工作表
    ws2 = wb2.active
    # 两个for循环变量整个excel的单元格内容
    # enumerate（）函数将用于一个可遍历的数据对象（如列表，元组或字符串）组合为一个索引序列，同时列出数据和数据下标，一般用在for循环中

    for i, row in enumerate(ws.iter_rows()):
        for j, cell in enumerate(row):
            # 写入新excel
            ws2.cell(row=i + 1, column=j + 1, value=cell.value)
            # print(ws2)
            ##设置列宽
            ws2.column_dimensions['A'].width = 20.0
            ws2.column_dimensions['B'].width = 60.0
            ws2.column_dimensions['C'].width = 30.0
            # 设置连续行行高：
            for r in range(1, -1):  # 注意，行和列的序数，都是从1开始
                ws.row_dimensions[r].height = 15
            # 设置新sheet的名称
            ws2.title = name
    wb2.save(name + ".xlsx")
# os.chdir(r'')


dir = input("请输入目标移动的路径: ")  ##定义路径
# print(dir)
for file in os.listdir(dir):
    ext = os.path.splitext(file)[1]  ##返回的是元组，后缀的索引是1， 用例列表[0],.xlsx[1]
    ext = ext[1:]  ##去掉”.“,通过切片，左开右闭
    if not os.path.isdir(f"{dir}/{ext}"):
        os.mkdir(f"{dir}/{ext}")
    source_path = f"{dir}/{file}"
    target_path = f"{dir}/{ext}/{file}"
    shutil.move(source_path, target_path)  ##文件移动

    print(file, ext)


"""
@Author   :xxxxx
@Contact  :1223242863@qq.com
@File     :zip.py
@Time     :2021/8/17 11:27 PM
@Software :Pycharm
@Copyright (c) 2021,All Rights Reserved.
"""

import os
import zipfile
from datetime import datetime

import datetime as datetime
from loguru import logger


def getZipDir(dirpath, outFullName):
    """
    压缩指定文件夹
    :param dirpath: 目标文件夹路径
    :param outFullName: 压缩文件保存路径+xxxx.zip
    :return: 无
    """
    zip = zipfile.ZipFile(outFullName, "w", zipfile.ZIP_DEFLATED)
    for path, dirnames, filenames in os.walk(dirpath):
        # 去掉目标跟路径，只对目标文件夹下边的文件及文件夹进行压缩
        fpath = path.replace(dirpath, '')

        for filename in filenames:
            zip.write(os.path.join(path, filename), os.path.join(fpath, filename))
    zip.close()
    logger.info("文件夹\"{0}\"已压缩为\"{1}\".".format(dirpath, outFullName))


if __name__ == "__main__":
    outFullName = './' + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + '_output.zip'
    getZipDir(dirpath="./",
              outFullName = outFullName)