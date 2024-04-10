import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
wb = Workbook()
ws = wb.active

if __name__ == '__main__':

    file_path = "D:\practice\github\py_project\\filter_table\\test_data_2024-03-22.xlsx"
    # rlt_path = "D:\practice\github\py_project\\filter_table\\test_data_2024-03-22_rlt.xlsx"
    # # # directory = 'D:\zhangyan7\Downloads\projectCase_2024-03-28-09-23-26_rlt'
    # if os.path.exists(rlt_path):
    #     print("文件已存在")
    # else:
    #     with open(rlt_path, 'w', encoding='utf-8') as f:
    #         print("文件创建成功")
    #         wb.save()
    # rlt_path = "D:\zhangyan7\Downloads\projectCase_2024-03-28-09-23-26_rlt.xlsx"

    df = pd.read_excel(file_path)
    valid_names = []
    invalid_names = []

    index = 0
    while index < len(df):
        value = df.iloc[index, 0]
        if str(value).startswith("语音_"):
            next_index = index + 1
            while next_index < len(df) and not df.iloc[next_index, 0].startswith("语音_"):
                next_index += 1

            if next_index - index > 2:
                valid_names.append(value)
            else:
                invalid_names.append(value)

            index = next_index
        else:
            index += 1

    with pd.ExcelWriter(file_path, engine="openpyxl", mode='a') as writer:
        pd.DataFrame(valid_names, columns=["有效数据"]).to_excel(writer, sheet_name="有效数据", index=False)
        pd.DataFrame(invalid_names, columns=["无效数据"]).to_excel(writer, sheet_name="无效数据", index=False)
        try:
            if "有效数据" in writer.sheetnames:
                writer.remove_sheet(writer.sheetnames["有效数据"])
                writer["有效数据"] = ['最新有效数据']
            if "无效数据" in writer.sheetnames:
                writer.remove_sheet(writer.sheetnames["无效数据"])
                writer["无效数据"] = ['最新无效数据']
            print("有效数据已写入")
            print("无效数据已写入")
        except Exception as e:
            print(f"写入Excel时出错：{e}")